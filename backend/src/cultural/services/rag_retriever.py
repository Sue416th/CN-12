import json

import os

from pathlib import Path

from typing import List, Dict, Any



# 支持读取 Excel 文件

try:

    import pandas as pd

    PANDAS_AVAILABLE = True

except ImportError:

    PANDAS_AVAILABLE = False





class RAGRetriever:

    """RAG检索器 - 从知识库中检索相关信息"""



    def __init__(self, knowledge_base_path: str = None):

        if knowledge_base_path is None:

            # 获取 backend 目录（src 的父目录的父目录）

            base_dir = Path(__file__).parent.parent.parent.parent

            knowledge_base_path = base_dir / "knowledge_base"

        self.knowledge_base_path = Path(knowledge_base_path)

        self.documents = self._load_knowledge_base()



    def _load_knowledge_base(self) -> List[Dict[str, Any]]:

        """加载所有知识库文件（JSON + Excel）"""

        documents = []



        if not self.knowledge_base_path.exists():

            return documents



        # 加载 JSON 文件

        for json_file in self.knowledge_base_path.glob("*.json"):

            try:

                with open(json_file, 'r', encoding='utf-8') as f:

                    data = json.load(f)

                    documents.append(data)

            except Exception as e:

                print(f"Error loading {json_file}: {e}")



        # 加载 Excel 文件

        if PANDAS_AVAILABLE:

            for excel_file in self.knowledge_base_path.glob("*.xlsx"):

                try:

                    # 使用 openpyxl 读取，解决编码问题

                    from openpyxl import load_workbook

                    wb = load_workbook(excel_file, data_only=True)

                    ws = wb.active



                    # 获取表头（第一行）

                    headers = []

                    for col in range(1, ws.max_column + 1):

                        headers.append(str(ws.cell(1, col).value))



                    # 读取每一行数据

                    for row_idx in range(2, ws.max_row + 1):

                        doc = {}

                        for col_idx, header in enumerate(headers, start=1):

                            cell_value = ws.cell(row_idx, col_idx).value

                            if cell_value:

                                doc[header] = str(cell_value)

                        if doc:

                            documents.append(doc)



                    print(f"Loaded {ws.max_row - 1} rows from {excel_file.name}")

                except Exception as e:

                    print(f"Error loading {excel_file}: {e}")



        return documents



    def retrieve(self, query: str, top_k: int = 3) -> List[Dict[str, Any]]:

        """根据查询检索相关文档"""

        query_lower = query.lower()

        results = []



        for doc in self.documents:

            relevance_score = self._calculate_relevance(query_lower, doc)

            if relevance_score > 0:

                results.append({

                    "document": doc,

                    "score": relevance_score

                })




        results.sort(key=lambda x: x["score"], reverse=True)

        return results[:top_k]



    def _calculate_relevance(self, query: str, doc: Dict[str, Any]) -> float:

        """计算文档与查询的相关性分数"""

        score = 0

        doc_text = json.dumps(doc, ensure_ascii=False)

        query_text = query



        # 目的地字段

        dest = doc.get('目的地', doc.get('attraction_name', ''))

        dest_text = str(dest) if dest else ""



        # 定义需要排除的通用问候语

        greetings = ["你好", "您好", "hello", "hi", "hey", "在吗", "在不在", "你好啊", "您好啊"]



        # 如果查询只是问候语，返回0分

        query_lower = query_text.lower()

        if query_lower.strip() in greetings or all(q in greetings for q in query_lower.split()):

            return 0



        # 中英关键词映射

        keyword_mapping = {

            # 英文到中文

            "lijiang": "丽江", "lijiang ancient town": "丽江古城",

            "west lake": "西湖", "west lake hangzhou": "西湖",

            "lingyin temple": "灵隐寺", "ling yin temple": "灵隐寺",

            "leifeng pagoda": "雷峰塔", "lei feng pagoda": "雷峰塔",

            "qiandao lake": "千岛湖", "thousand island lake": "千岛湖",

            "broken bridge": "断桥", "su causeway": "苏堤",

            "phoenix ancient town": "凤凰古城", "phoenix": "凤凰古城",

            "zhangjiajie": "张家界",

            "wuzhen": "乌镇",

            "yulong snow mountain": "玉龙雪山", "jade dragon snow mountain": "玉龙雪山",

            "shuhe ancient town": "束河古镇", "shuhe": "束河古镇",

            "luguhu": "泸沽湖", "lugu lake": "泸沽湖",

            "shangri-la": "香格里拉",

            "dali": "大理", "erhai": "洱海", "erhai lake": "洱海",

            "xintang": "西塘", "nanxun": "南浔",

            "jiuzhaigou": "九寨沟",

 "emeishan": "峨眉山", "mount emei": "峨眉山",

            "leshan giant buddha": "乐山大佛",

            "accommodation": "住宿", "hotel": "住宿", "guesthouse": "住宿",

            "food": "美食", "restaurant": "美食",

            "ticket": "门票",

            "history": "历史", "historical": "历史",

            "culture": "文化", "cultural": "文化",

            "legend": "传说",

            "transportation": "交通", "transport": "交通",

            "recommend": "推荐", "recommendation": "推荐",

            "attraction": "景点", "spot": "景点",

        }



        # 先处理英文关键词，转换为中文

        query_lower = query_text.lower()

        for eng_key, chi_key in keyword_mapping.items():

            if eng_key in query_lower:

                query_text = query_text + " " + chi_key



        # 核心关键词 - 精确匹配

        core_keywords = {

            "西湖": 15, "west lake": 15,

            "灵隐寺": 15, "lingyin temple": 15,

            "雷峰塔": 15, "leifeng pagoda": 15,

            "千岛湖": 15, "断桥": 15, "苏堤": 15,

            "凤凰古城": 15, "张家界": 15, "乌镇": 15,

            "丽江": 15, "丽江古城": 15,

            "玉龙雪山": 15, "束河古镇": 15, "泸沽湖": 15,

            "香格里拉": 15, "大理": 12, "洱海": 12,

            "西塘": 12, "南浔": 12,

            "九寨沟": 15, "峨眉山": 12, "乐山大佛": 12,

        }



        # 检查核心关键词（标题或目的地）

        for keyword, weight in core_keywords.items():

            if keyword in query_text:

                if keyword in doc_text or (dest and keyword in dest):

                    score += weight



        # 动态匹配：查询中的每个2+字符的词是否在目的地字段中

        if dest_text:

            # 提取查询中所有2-6个连续字符的词

            for i in range(len(query_text)):

                for length in range(6, 1, -1):

                    if i + length <= len(query_text):

                        word = query_text[i:i+length]

                        # 跳过纯数字和问候语

                        if len(word) >= 2 and not word.isdigit() and word not in greetings:

                            if word in dest_text:

                                score += 15

                                break



        # 内容关键词匹配

        content_keywords = {

            "历史": 8, "文化": 8, "传说": 8,

            "门票": 5, "开放时间": 5,

            "交通": 8, "住宿": 8, "美食": 8, "推荐": 5,

            "景点": 5, "旅游": 5, "旅行": 5,

        }



        for keyword, weight in content_keywords.items():

            if keyword in query_text and keyword in doc_text:

                score += weight



        return score



    def format_retrieved_context(self, results: List[Dict[str, Any]]) -> str:

        """格式化检索结果为上下文字符串"""

        if not results:

            return ""



        context_parts = []

        for i, result in enumerate(results, 1):

            doc = result["document"]



            # 检测文档类型并格式化

            if "attraction_name" in doc:

                # JSON 格式（旧格式）

                context_parts.append(f"\n【相关景点{i}】: {doc.get('attraction_name', 'Unknown')}\n")

                context_parts.append(f"位置: {doc.get('location', 'Unknown')}\n")

                context_parts.append(f"历史时期: {doc.get('historical_period', 'Unknown')}\n")

                context_parts.append(f"文化意义: {doc.get('cultural_significance', 'Unknown')}\n")



                if 'highlights' in doc:

                    highlights = doc['highlights']

                    if isinstance(highlights, list):

                        context_parts.append(f"主要景点: {', '.join(highlights)}\n")



                if 'legend' in doc:

                    legend = doc['legend']

                    if isinstance(legend, dict):

                        context_parts.append(f"传说: {legend.get('title', '')} - {legend.get('description', '')}\n")



                if 'history' in doc:

                    history = doc['history']

                    if isinstance(history, list):

                        context_parts.append("历史事件:\n")

                        for h in history:

                            context_parts.append(f"  - {h.get('period', '')}: {h.get('event', '')}\n")



                if 'visiting_info' in doc:

                    info = doc['visiting_info']

                    if isinstance(info, dict):

                        context_parts.append(f"游览信息: {json.dumps(info, ensure_ascii=False)}\n")

            else:

                # Excel 格式（新格式）

                context_parts.append(f"\n【相关景点{i}】: {doc.get('目的地', 'Unknown')}\n")



                # 遍历所有字段

                for key, value in doc.items():

                    if key and value:

                        context_parts.append(f"{key}: {value}\n")



        return "".join(context_parts)
